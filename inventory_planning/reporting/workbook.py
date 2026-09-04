"""
The run, as one workbook a planner opens.

A run used to leave sixteen CSVs and two other files in a folder. Every one of them was
written for a reason, and together they were unusable: the question a planner arrives
with — *what do I buy, and is this item in trouble* — is answered by joining four of
them, and nobody joins four CSVs in a review meeting. The files were also named for the
stage that produced them rather than for the question they answer, so knowing which to
open was itself institutional knowledge.

Five sheets, each the answer to one question:

    Forecast     what it sold, what it will sell, and how much the model is worth
    Parameters   every planning parameter per item, and where each value came from
    Purchase     what to order, what to pull in, what to push out
    Inventory    what is on the shelf, how long it lasts, and against what policy
    S&IOP        projected on-hand per item per month, with the purchase behind it

The S&IOP sheet is per-item rather than a rollup. The rollup existed and was the wrong
shape: a gap by product line is a number to be explained, not acted on, and the line it
sums is exactly what a planner needs to see. Summed, this sheet *is* the rollup — the
totals fall out of it and cannot disagree with it.

## Column order, and why nothing is dropped

Each sheet leads with a curated, ordered set: identity, then the two or three figures
the sheet exists for, then their provenance. Every remaining column of the source frame
follows to the right, unordered.

That is deliberate. Curating the front is what makes the sheet readable; dropping the
rest would make this a different, lossier report than the CSVs it replaces, and the
column somebody's spreadsheet depends on is never the one you would have guessed. The
first screen is for reading; everything to the right of it is for filtering.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd

MONEY = "#,##0"
QTY = "#,##0"
DAYS = "#,##0.0"
RATE = "0%"
RATIO = "0.00"

# Sheet name -> the columns that lead it, in order. A column named here and absent from
# the frame is skipped rather than raising: the pipeline runs on extracts of different
# richness, and a workbook that refuses to build because one optional measure was not
# supplied is worse than one with a shorter first screen.
LEAD_COLUMNS: Dict[str, Sequence[Tuple[str, Optional[str]]]] = {
    "Forecast": [
        ("sku", None), ("description", None), ("product_family", None),
        ("model_used", None), ("selected_by", None),
        ("vs_naive", RATIO), ("backtest_mase", RATIO), ("backtest_mape", RATE),
        ("expected_order_size", QTY), ("expected_interval", RATIO),
    ],
    "Parameters": [
        ("sku", None), ("description", None), ("product_family", None),
        ("business_unit", None), ("abc_class", None), ("stocking_class", None),
        ("demand_pattern", None), ("replenishment_method", None),
        ("review_period_days", DAYS), ("service_level", RATE),
        ("lead_time_days", DAYS), ("lt_sigma_days", DAYS), ("lt_samples", QTY),
        ("unit_cost", MONEY), ("min_order_qty", QTY), ("order_multiple", QTY),
        ("demand_mean", QTY), ("demand_sigma", QTY),
        ("lead_time_days_source", None), ("unit_cost_source", None),
        ("min_order_qty_source", None), ("order_multiple_source", None),
        ("product_family_source", None), ("sigma_source", None),
        ("applied_rules", None),
    ],
    "Purchase": [
        ("sku", None), ("description", None), ("product_family", None),
        ("recommended_action", None),
        ("suggested_po_qty", QTY), ("pushout_open_po_qty", QTY),
        ("mto_order_by", None), ("mto_order_status", None),
        ("net_requirement", QTY), ("order_quantity", QTY), ("order_lot", QTY),
        ("reorder_point", QTY), ("order_up_to", QTY),
        ("available_supply", QTY), ("period_demand", QTY), ("demand_driver", None),
        ("forecast_next_period", QTY), ("backlog_due_qty", QTY),
        ("days_to_next_arrival", DAYS), ("supply_gap", QTY), ("supply_gap_days", DAYS),
        ("wma_lead_time_days", DAYS),
    ],
    "Inventory": [
        ("sku", None), ("description", None), ("product_family", None),
        ("abc_class", None), ("stocking_class", None), ("inventory_status", None),
        ("actual_qty", QTY), ("actual_value", MONEY),
        ("actual_dioh", DAYS), ("days_of_supply", DAYS), ("on_hand_cover_days", DAYS),
        ("should_be_qty", QTY), ("should_be_value", MONEY), ("should_be_dioh", DAYS),
        ("gap_qty", QTY), ("gap_value", MONEY), ("coverage_ratio", RATIO),
        ("cycle_qty", QTY), ("safety_qty", QTY), ("pipeline_qty", QTY),
        ("committed_qty", QTY), ("is_non_stocking", None),
        ("qty_on_hand", QTY), ("qty_in_transit", QTY), ("unit_cost", MONEY),
        ("last_sale", None),
    ],
    "S&IOP": [
        ("sku", None), ("description", None), ("product_family", None),
        ("period", None),
        ("demand_qty", QTY), ("supply_qty", QTY), ("closing_qty", QTY),
        ("safety_stock", QTY), ("gap_qty", QTY),
        ("demand_cogs", MONEY), ("supply_value", MONEY), ("closing_value", MONEY),
        ("gap_value", MONEY),
    ],
}

# What each sheet is for, in the file rather than in the message that sent it. A
# workbook that circulates for three weeks outlives every explanation given with it.
SHEET_NOTES: Dict[str, List[str]] = {
    "Forecast": [
        "One row per item: history to the left, forecast to the right, in the same units.",
        "`model_used` is the winner of a backtest against every other model, including "
        "doing nothing. `vs_naive` is its error over the error of simply repeating last "
        "month — 1.00 means the model added nothing.",
        "`selected_by` says which metric decided. MASE where the item has months with no "
        "demand, MAPE only where it has none — a percentage cannot be computed on a zero, "
        "and scoring only the months that sold inflates the forecast of an intermittent item.",
        "`expected_order_size` and `expected_interval` are the lump behind the rate. An "
        "item forecast at 33/month that orders 100 once a quarter shows 100 and 3.0 here; "
        "the flat rate is right for safety stock and misleading as a plan.",
    ],
    "Parameters": [
        "Every planning parameter in force, per item, with the source of each value.",
        "`*_source` columns say where the number came from: measured from transactions, "
        "the ERP item master, the planning master, or a config default. Where two sources "
        "disagreed, the one shown is the one used.",
        "`applied_rules` lists the planner rules that set this item's review period and "
        "service level.",
    ],
    "Purchase": [
        "What to do about each item this cycle.",
        "`suggested_po_qty` is a new order. `pushout_open_po_qty` is an existing order to "
        "delay or cancel. `mto_order_by` is the date an order-on-demand item's purchase "
        "order has to be placed to meet a customer date already on the book.",
        "`period_demand` is max(forecast, firm backlog), not their sum — the two are "
        "estimates of one demand, and adding them buys it twice. `demand_driver` says "
        "which one won.",
    ],
    "Inventory": [
        "The position per item, and what policy says it should be.",
        "`should_be_qty` splits into cycle + safety + pipeline. An order-on-demand item "
        "holds none of those, but is still sized against `committed_qty` — the firm order "
        "book it has to cover. Stock above that is genuine excess; stock below it is a "
        "shortfall against a customer commitment.",
        "`gap_value` is actual minus should-be. Positive is capital tied up, negative is "
        "a service risk. They do not net off: the units are not interchangeable.",
        "DIOH is days of *cover* where no ageing date was supplied — a fast item restocked "
        "yesterday can show a year of cover, and an old part with one order shows almost none.",
    ],
    "S&IOP": [
        "Projected on-hand per item per month, and the purchase behind it.",
        "`closing_qty` is the projected position at the end of the period. A period is "
        "short when it falls below `safety_stock`, which is what `gap_qty` measures — not "
        "when the shelf is bare.",
        "Summed by product line or by period, this sheet is the S&IOP rollup. The totals "
        "come out of these rows, so the two cannot disagree.",
        "Amounts are at cost. Items with no unit cost carry quantities and a blank amount, "
        "so they cannot sum into a total unnoticed.",
    ],
}


def _order_columns(frame: pd.DataFrame, sheet: str) -> Tuple[pd.DataFrame, Dict[str, str]]:
    """Curated columns first, in order; everything else after, unchanged."""
    lead = [(c, fmt) for c, fmt in LEAD_COLUMNS.get(sheet, []) if c in frame.columns]
    lead_names = [c for c, _ in lead]
    rest = [c for c in frame.columns if c not in lead_names]
    formats = {c: fmt for c, fmt in lead if fmt}
    return frame[lead_names + rest], formats


def _sku_dimensions(policy: Dict[str, Any], results: Dict[str, Any]) -> pd.DataFrame:
    """
    Description and product line per SKU, to carry onto every sheet.

    A sheet keyed only on a material number is one a planner has to decode against
    another sheet, which is the failure this workbook exists to end.
    """
    wanted = ["sku", "description", "product_family", "business_unit"]
    for source in (policy.get("sku_attributes"), results.get("sku_attributes"),
                   getattr(policy.get("should_be"), "frame", None)):
        if isinstance(source, pd.DataFrame) and "sku" in source.columns:
            cols = [c for c in wanted if c in source.columns]
            if len(cols) > 1:
                return source[cols].drop_duplicates("sku")
    return pd.DataFrame(columns=["sku"])


def _attach(frame: pd.DataFrame, dims: pd.DataFrame) -> pd.DataFrame:
    if frame is None or not len(frame) or "sku" not in frame.columns or not len(dims):
        return frame
    add = [c for c in dims.columns if c != "sku" and c not in frame.columns]
    if not add:
        return frame
    out = frame.merge(dims[["sku"] + add], on="sku", how="left")
    return out


def _merge_extra(frame: pd.DataFrame, other, suffix: str) -> pd.DataFrame:
    """
    Fold another per-SKU frame in, keeping only what the sheet does not already carry.

    Same-named columns are dropped rather than suffixed. Two copies of `unit_cost` on
    one row is not extra information — it is a question about which one is real, on
    every row, for a reader who cannot answer it.
    """
    other = getattr(other, "frame", other)
    if frame is None or not isinstance(other, pd.DataFrame) or not len(other):
        return frame
    if "sku" not in other.columns:
        return frame
    add = [c for c in other.columns if c != "sku" and c not in frame.columns]
    if not add:
        return frame
    right = other[["sku"] + add].drop_duplicates("sku")
    if suffix:
        right = right.rename(columns={c: f"{c}{suffix}" for c in add})
    return frame.merge(right, on="sku", how="left")


def collect_sheets(results: Dict[str, Any],
                   policy: Dict[str, Any] = None) -> Dict[str, pd.DataFrame]:
    """
    The five frames, already joined to their dimensions and ordered.

    Separated from writing so the same selection can be tested without a file, and so a
    caller that wants the frames for something else does not have to read an xlsx back.
    """
    policy = policy or {}
    dims = _sku_dimensions(policy, results)
    sheets: Dict[str, pd.DataFrame] = {}

    forecast = results.get("forecast_sheet")
    if forecast is None or not len(forecast):
        forecast = results.get("forecast_detail")
    if forecast is not None and len(forecast):
        frame = forecast.reset_index() if forecast.index.name == "sku" else forecast
        accuracy = policy.get("forecast_accuracy") or results.get("forecast_accuracy")
        # How the last plan actually scored, beside the one being published now. Kept
        # on the same row because the question they answer together — is this model
        # earning its place on this item — cannot be asked across two files.
        frame = _merge_extra(frame, getattr(accuracy, "by_sku", None), "_prior_plan")
        frame = _merge_extra(frame, getattr(accuracy, "adjustments", None), "_review")
        sheets["Forecast"] = _attach(frame, dims)

    should_be = getattr(policy.get("should_be"), "frame", None)
    parameters = should_be if should_be is not None else getattr(
        results.get("parameters"), "frame", None)
    if parameters is not None and len(parameters):
        frame = _attach(parameters, dims)
        frame = _merge_extra(frame, results.get("policy_profile"), "")
        frame = _merge_extra(frame, policy.get("suggestions"), "_suggested")
        sheets["Parameters"] = frame

    recommendations = results.get("recommendations")
    if recommendations is not None and len(recommendations):
        frame = _attach(recommendations, dims)
        realization = policy.get("backlog_realization") or results.get("backlog_realization")
        frame = _merge_extra(frame, getattr(realization, "per_sku", None), "_realization")
        sheets["Purchase"] = frame

    inventory = should_be
    projection = results.get("projection")
    if inventory is not None and len(inventory):
        if projection is not None and len(projection):
            extra = [c for c in ("days_of_supply", "on_hand_cover_days",
                                 "inventory_status", "effective_position")
                     if c in projection.columns and c not in inventory.columns]
            if extra:
                inventory = inventory.merge(projection[["sku"] + extra], on="sku",
                                            how="left")
        sheets["Inventory"] = _attach(inventory, dims)

    siop = policy.get("siop") or results.get("siop")
    per_sku = getattr(siop, "per_sku", None)
    if per_sku is not None and len(per_sku):
        sheets["S&IOP"] = _attach(per_sku, dims)

    return {name: _order_columns(frame, name)[0] for name, frame in sheets.items()}


def build_workbook(path, results: Dict[str, Any], policy: Dict[str, Any] = None,
                   currency: str = "USD") -> Optional[Path]:
    """
    Write the run as one workbook. Returns the path, or None if there was nothing to write.
    """
    sheets = collect_sheets(results, policy)
    if not sheets:
        return None

    path = Path(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
            _format_sheet(writer.book[name], frame, name)

        notes = []
        for name in sheets:
            notes.append({"sheet": name, "note": ""})
            for line in SHEET_NOTES.get(name, []):
                notes.append({"sheet": "", "note": line})
        notes.append({"sheet": "", "note": ""})
        notes.append({"sheet": "Amounts", "note":
                      f"Every amount in this workbook is in {currency}."})
        notes.append({"sheet": "", "note":
                      "Columns to the right of the first screen are the rest of the "
                      "source data, unordered. Nothing has been dropped."})
        pd.DataFrame(notes).to_excel(writer, sheet_name="How to read this", index=False)
        _format_sheet(writer.book["How to read this"],
                      pd.DataFrame(notes), "How to read this", widths=(14, 110))

    return path


def _format_sheet(worksheet, frame: pd.DataFrame, sheet: str,
                  widths: Tuple[int, int] = None) -> None:
    """Freeze the header, filter every column, and give the numbers a readable format."""
    from openpyxl.utils import get_column_letter

    worksheet.freeze_panes = "A2"
    if len(frame.columns) and len(frame):
        worksheet.auto_filter.ref = worksheet.dimensions

    if widths:
        for i, width in enumerate(widths, start=1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        from openpyxl.styles import Alignment

        for row in worksheet.iter_rows(min_row=1, min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        return

    _, formats = _order_columns(frame, sheet)
    for i, column in enumerate(frame.columns, start=1):
        letter = get_column_letter(i)
        header = len(str(column))
        sample = frame[column].head(200).astype(str).str.len().max()
        worksheet.column_dimensions[letter].width = min(
            42, max(10, header + 2, int(sample) + 2 if pd.notna(sample) else 10))
        fmt = formats.get(column)
        if not fmt:
            continue
        for row in worksheet.iter_rows(min_row=2, min_col=i, max_col=i):
            for cell in row:
                cell.number_format = fmt
