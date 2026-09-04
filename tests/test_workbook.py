"""
The run as one workbook.

A run used to leave sixteen CSVs named for the stage that produced them. The question a
planner arrives with — what do I buy, and is this item in trouble — was answered by
joining four of them, and nobody joins four CSVs in a review meeting.

What these hold: the five sheets exist and are keyed so a planner can read a row without
decoding it against another file; the curated columns lead and nothing is dropped behind
them; and the S&IOP sheet sums to the rollup it replaced, because a detail sheet that
disagrees with its own total is worse than the two files it merged.
"""

import glob
import io
import contextlib
import sys
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).parents[1]))

from inventory_planning.orchestrator import InventoryPlanner
from inventory_planning.reporting.workbook import (
    LEAD_COLUMNS, build_workbook, collect_sheets,
)

SAMPLE = Path(__file__).parents[1] / "sample_data"


@pytest.fixture(scope="module")
def run(tmp_path_factory):
    out = tmp_path_factory.mktemp("workbook")
    with contextlib.redirect_stdout(io.StringIO()):
        planner = InventoryPlanner(output_dir=out, interactive=False)
        results = planner.run_planning(
            **planner.load_all(sorted(glob.glob(str(SAMPLE / "*.csv")))))
        policy = planner.run_policy_analysis(
            results, inventory_df=results.get("inventory_consolidated"))
    return planner, results, policy, out


class TestTheFiveSheets:

    def test_every_sheet_is_built(self, run):
        _, results, policy, _ = run
        assert set(collect_sheets(results, policy)) == {
            "Forecast", "Parameters", "Purchase", "Inventory", "S&IOP"}

    def test_each_one_is_keyed_on_the_item(self, run):
        _, results, policy, _ = run
        for name, frame in collect_sheets(results, policy).items():
            assert frame.columns[0] == "sku", f"{name} does not lead with the item"

    def test_a_row_can_be_read_without_another_file(self, run):
        """A sheet keyed only on a material number is one a planner has to decode."""
        _, results, policy, _ = run
        for name, frame in collect_sheets(results, policy).items():
            assert "product_family" in frame.columns, name

    def test_the_curated_columns_lead(self, run):
        _, results, policy, _ = run
        for name, frame in collect_sheets(results, policy).items():
            wanted = [c for c, _ in LEAD_COLUMNS[name] if c in frame.columns]
            assert list(frame.columns[:len(wanted)]) == wanted

    def test_nothing_behind_them_is_dropped(self, run):
        """
        Curating the front is what makes a sheet readable. Dropping the rest would make
        this a lossier report than the CSVs it replaces, and the column somebody's
        spreadsheet depends on is never the one you would have guessed.
        """
        _, results, policy, _ = run
        sheets = collect_sheets(results, policy)
        for source, sheet in ((results["recommendations"], "Purchase"),
                              (policy["should_be"].frame, "Parameters"),
                              (policy["siop"].per_sku, "S&IOP")):
            missing = set(source.columns) - set(sheets[sheet].columns)
            assert not missing, f"{sheet} dropped {sorted(missing)}"


class TestTheSiopSheetIsTheRollup:
    """
    The by-family rollup existed and was the wrong shape: a gap by product line is a
    number to be explained, not acted on. Per item, summed, it *is* the rollup — and a
    detail sheet that disagrees with its own total is worse than the two files it
    merged.
    """

    def test_it_is_per_item_per_period(self, run):
        _, results, policy, _ = run
        sheet = collect_sheets(results, policy)["S&IOP"]
        assert {"sku", "period", "closing_qty"} <= set(sheet.columns)
        assert len(sheet) == len(sheet.drop_duplicates(["sku", "period"]))

    def test_it_sums_to_the_period_totals(self, run):
        _, results, policy, _ = run
        sheet = collect_sheets(results, policy)["S&IOP"]
        by_period = policy["siop"].by_period.set_index("period")
        rolled = sheet.groupby("period")[["demand_cogs", "closing_value"]].sum()
        for period, row in rolled.iterrows():
            assert row["demand_cogs"] == pytest.approx(
                by_period.loc[period, "demand_cogs"], rel=1e-6)
            assert row["closing_value"] == pytest.approx(
                by_period.loc[period, "closing_value"], rel=1e-6)

    def test_the_purchase_behind_the_position_is_on_the_row(self, run):
        """Projected on-hand without the buy that produces it is not actionable."""
        _, results, policy, _ = run
        sheet = collect_sheets(results, policy)["S&IOP"]
        assert {"supply_qty", "supply_value"} <= set(sheet.columns)


class TestTheFileItself:

    def test_it_writes_and_carries_the_notes(self, run):
        import openpyxl

        _, results, policy, out = run
        path = build_workbook(out / "wb.xlsx", results, policy, currency="CNY")
        book = openpyxl.load_workbook(path)
        assert book.sheetnames == ["Forecast", "Parameters", "Purchase", "Inventory",
                                   "S&IOP", "How to read this"]
        notes = pd.read_excel(path, sheet_name="How to read this")
        assert notes["note"].str.contains("in CNY").any(), \
            "a workbook that does not name its currency is how one gets read as another"

    def test_every_sheet_freezes_its_header(self, run):
        import openpyxl

        _, results, policy, out = run
        book = openpyxl.load_workbook(build_workbook(out / "wb2.xlsx", results, policy))
        assert all(ws.freeze_panes == "A2" for ws in book.worksheets)

    def test_a_run_with_no_policy_stage_still_writes_what_it_has(self, run):
        """A run that stops at the gate should leave something readable, not a folder."""
        _, results, _, out = run
        sheets = collect_sheets(results, None)
        assert "Forecast" in sheets and "Purchase" in sheets
        assert build_workbook(out / "wb3.xlsx", results, None) is not None

    def test_nothing_to_report_is_not_an_empty_file(self, run):
        _, _, _, out = run
        assert build_workbook(out / "wb4.xlsx", {}, {}) is None
        assert not (out / "wb4.xlsx").exists()

    def test_the_run_writes_it_and_records_it(self, run):
        planner, _, _, out = run
        path = out / f"planning_{planner.run.run_id}.xlsx"
        assert path.exists()
        assert any(Path(o.name).name == path.name for o in planner.run.outputs), \
            "an output the manifest does not name has no provenance"


class TestOneNumberPerRow:

    def test_a_folded_frame_does_not_bring_a_second_copy_of_a_column(self, run):
        """
        Two `unit_cost` columns on one row is not extra information — it is a question
        about which one is real, on every row, for a reader who cannot answer it.
        """
        _, results, policy, _ = run
        for name, frame in collect_sheets(results, policy).items():
            duplicated = [c for c in frame.columns if list(frame.columns).count(c) > 1]
            assert not duplicated, f"{name} carries {duplicated} twice"
